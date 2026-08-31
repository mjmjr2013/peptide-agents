"""
The labeling manifest Jason's crew works from — 2026-08-31.

What these tests protect, in order of how much it would cost to get wrong:

  1. STRENGTH IS NEVER AMBIGUOUS. Mislabelling 10mg as 100mg is the failure this
     manifest exists to prevent, so the strength gets its own column and every
     row must carry it.
  2. A SKU WITH NO STICKER IS CALLED OUT, never silently blank. A blank cell at a
     bench reads as "no sticker needed"; the crew must be stopped instead.
  3. PACKAGES MATCH WHAT SHIPS. The crew packs parcels, not orders, so the sheet
     must break an order into the same packages core/shipping does.
  4. THE TWO TABS ARE DISJOINT AND COMPLETE. An order that falls into neither is
     an order nobody works on.

Run:  python3 -m pytest tests/test_manifest.py -q
"""
from __future__ import annotations
import io
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from core import catalog, manifest, shipping                     # noqa: E402


# ── Fixtures ─────────────────────────────────────────────────────────────────

@pytest.fixture
def stickers(tmp_path, monkeypatch):
    """Three SKUs with artwork on file; everything else deliberately has none.

    Points LABEL_DIR at a temp folder so these tests describe manifest behavior
    rather than whatever artwork happens to be installed — the real files are
    covered by test_labels.py.
    """
    from PIL import Image
    monkeypatch.setattr(catalog, "LABEL_DIR", tmp_path)
    for sku in ("RT10", "RT100", "BAC10"):
        Image.new("RGB", (660, 270), (31, 42, 68)).save(tmp_path / f"{sku}.png")
    # BAC10 deliberately has NO recorded wording, to exercise the fallback.
    monkeypatch.setattr(catalog, "SKU_LABEL_TEXT",
                        {"RT10": "GLP-3 RT 10mg", "RT100": "GLP-3 RT 100mg"})
    return {"RT10": "RT10.png", "RT100": "RT100.png", "BAC10": "BAC10.png"}


def order(ref, items, tracked=False, photo=False, name="Jane Doe"):
    return {"id": ref, "fields": {
        "order_ref": ref, "ship_name": name, "address_line1": "12 Main St",
        "city": "Provo", "state_province": "UT", "postal_code": "84604",
        "country": "USA", "ship_phone": "+15551230000",
        "tracking_sent": tracked, "vial_photo_sent": photo, "_items": items}}


def fetch_items(o):
    return [{"fields": i} for i in o["fields"]["_items"]]


def line(sku, product, spec, kits):
    return {"supplier_sku": sku, "product": product, "spec": spec, "kits": kits}


def sheet_values(data: bytes, index: int = 0):
    wb = load_workbook(io.BytesIO(data))
    ws = wb[wb.sheetnames[index]]
    return ws, [[c.value for c in row] for row in ws.iter_rows()]


# ── The two tabs ─────────────────────────────────────────────────────────────

def test_split_by_stage_is_disjoint_and_complete():
    orders = [
        order("A", [], tracked=False, photo=False),
        order("B", [], tracked=True, photo=False),
        order("C", [], tracked=False, photo=True),
        order("D", [], tracked=True, photo=True),      # nothing outstanding
    ]
    new, photo = manifest.split_by_stage(orders)
    new_refs = {o["fields"]["order_ref"] for o in new}
    photo_refs = {o["fields"]["order_ref"] for o in photo}
    assert new_refs == {"A", "C"}
    assert photo_refs == {"B"}
    assert new_refs & photo_refs == set(), "an order must not appear on both tabs"


def test_workbook_has_exactly_the_two_expected_tabs(stickers):
    data = manifest.build_labeling_manifest([], [], "today", fetch_items)
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["1 · Label & Ship", "2 · Photo Before Ship"]


def test_empty_tabs_say_so_rather_than_looking_broken(stickers):
    data = manifest.build_labeling_manifest([], [], "today", fetch_items)
    for i in (0, 1):
        _ws, rows = sheet_values(data, i)
        flat = " ".join(str(c) for r in rows for c in r if c)
        assert "Nothing new to label" in flat or "waiting on a photo" in flat


def test_orders_land_on_the_right_tab(stickers):
    new = [order("NL-NEW", [line("RT10", "Retatrutide", "10mg x10", 1)])]
    photo = [order("NL-PIC", [line("RT100", "Retatrutide", "100mg x10", 1)], tracked=True)]
    data = manifest.build_labeling_manifest(new, photo, "today", fetch_items)
    _ws, tab1 = sheet_values(data, 0)
    _ws, tab2 = sheet_values(data, 1)
    assert any("NL-NEW" in str(c) for r in tab1 for c in r if c)
    assert not any("NL-PIC" in str(c) for r in tab1 for c in r if c)
    assert any("NL-PIC" in str(c) for r in tab2 for c in r if c)


# ── Strength ─────────────────────────────────────────────────────────────────

def test_strength_has_its_own_column_on_every_item_row(stickers):
    orders = [order("NL-1", [line("RT10", "Retatrutide", "10mg x10", 3),
                             line("RT100", "Retatrutide", "100mg x10", 2)])]
    data = manifest.build_labeling_manifest(orders, [], "today", fetch_items)
    _ws, rows = sheet_values(data, 0)
    items = {r[0]: r for r in rows if r[0] in ("RT10", "RT100")}
    assert items["RT10"][2] == "10mg"
    assert items["RT100"][2] == "100mg"
    assert items["RT10"][3] == 3 and items["RT100"][3] == 2


def test_the_two_strengths_are_never_the_same_string(stickers):
    """10mg and 100mg differ by one character in a product name. In their own
    column they cannot be confused — this asserts they never collapse."""
    orders = [order("NL-1", [line("RT10", "Retatrutide", "10mg x10", 1),
                             line("RT100", "Retatrutide", "100mg x10", 1)])]
    data = manifest.build_labeling_manifest(orders, [], "today", fetch_items)
    _ws, rows = sheet_values(data, 0)
    strengths = [r[2] for r in rows if r[0] in ("RT10", "RT100")]
    assert strengths == ["10mg", "100mg"] or strengths == ["100mg", "10mg"]
    assert len(set(strengths)) == 2


def test_every_item_row_carries_a_strength(stickers):
    lines = [line(i.sku, i.product, i.spec, 1) for i in list(catalog.ITEMS)[:40]]
    data = manifest.build_labeling_manifest(
        [order("NL-BIG", lines)], [], "today", fetch_items)
    _ws, rows = sheet_values(data, 0)
    skus = {i.sku for i in catalog.ITEMS}
    blank = [r[0] for r in rows if r[0] in skus and not r[2]]
    assert blank == [], f"rows with no strength: {blank}"


# ── Stickers ─────────────────────────────────────────────────────────────────

def test_the_sticker_image_is_embedded_when_one_is_on_file(stickers):
    orders = [order("NL-1", [line("RT10", "Retatrutide", "10mg x10", 1)])]
    data = manifest.build_labeling_manifest(orders, [], "today", fetch_items)
    wb = load_workbook(io.BytesIO(data))
    assert len(wb["1 · Label & Ship"]._images) == 1


def test_a_sku_with_no_sticker_is_called_out_not_left_blank(stickers):
    """A blank cell at a bench reads as 'no sticker needed'. It must read as stop."""
    orders = [order("NL-1", [line("SM10", "Semaglutide", "10mg x10", 2)])]
    data = manifest.build_labeling_manifest(orders, [], "today", fetch_items)
    ws, rows = sheet_values(data, 0)
    row = next(r for r in rows if r[0] == "SM10")
    assert row[4] and "NO STICKER ON FILE" in str(row[4])
    assert len(ws._images) == 0
    cell = next(c for r in ws.iter_rows() for c in r
                if c.value and "NO STICKER" in str(c.value))
    assert cell.font.color.rgb.endswith(manifest.RED), "the warning must be red"


def test_label_text_is_the_sticker_wording_when_we_have_it(stickers):
    orders = [order("NL-1", [line("RT10", "Retatrutide", "10mg x10", 1)])]
    data = manifest.build_labeling_manifest(orders, [], "today", fetch_items)
    _ws, rows = sheet_values(data, 0)
    assert next(r for r in rows if r[0] == "RT10")[1] == "GLP-3 RT 10mg"


def test_label_text_falls_back_to_the_catalog_name(stickers):
    """BAC10 has artwork but no recorded sticker wording — the row is still usable."""
    orders = [order("NL-1", [line("BAC10", "Bacteriostatic Water", "10ml x10", 1)])]
    data = manifest.build_labeling_manifest(orders, [], "today", fetch_items)
    _ws, rows = sheet_values(data, 0)
    assert next(r for r in rows if r[0] == "BAC10")[1] == "Bacteriostatic Water 10ml"


# ── Packages ─────────────────────────────────────────────────────────────────

def test_a_heavy_order_is_broken_into_the_packages_that_actually_ship(stickers):
    orders = [order("NL-HEAVY", [line("RT100", "Retatrutide", "100mg x10", 40)])]
    data = manifest.build_labeling_manifest(orders, [], "today", fetch_items)
    _ws, rows = sheet_values(data, 0)
    banners = [str(r[0]) for r in rows if r[0] and "PACKAGE" in str(r[0])]
    assert len(banners) == 2, banners
    assert "PACKAGE 1 of 2" in banners[0] and "PACKAGE 2 of 2" in banners[1]
    assert all("20 kits" in b for b in banners), "the split must be balanced"


def test_every_package_banner_names_its_own_order(stickers):
    """A big order runs over a page break; a package header stranded at the top of
    a page with no order above it is a parcel nobody can identify."""
    orders = [order("NL-HEAVY", [line("RT100", "Retatrutide", "100mg x10", 40)]),
              order("NL-SMALL", [line("RT10", "Retatrutide", "10mg x10", 1)])]
    data = manifest.build_labeling_manifest(orders, [], "today", fetch_items)
    _ws, rows = sheet_values(data, 0)
    for r in rows:
        if r[0] and "PACKAGE" in str(r[0]):
            assert "NL-" in str(r[0]), f"package banner without an order ref: {r[0]}"


def test_water_ships_as_one_package_and_says_why(stickers):
    orders = [order("NL-WATER", [line("BAC10", "Bacteriostatic Water", "10ml x10", 84)])]
    data = manifest.build_labeling_manifest(orders, [], "today", fetch_items)
    _ws, rows = sheet_values(data, 0)
    banners = [str(r[0]) for r in rows if r[0] and "PACKAGE" in str(r[0])]
    assert len(banners) == 1, "bac water is exempt from the 2 kg cap"
    assert "84 kits" in banners[0] and "no weight cap" in banners[0]


def test_kit_counts_on_the_sheet_add_up_to_the_order(stickers):
    lines = [line("RT100", "Retatrutide", "100mg x10", 12),
             line("BAC10", "Bacteriostatic Water", "10ml x10", 9),
             line("RT10", "Retatrutide", "10mg x10", 5)]
    data = manifest.build_labeling_manifest([order("NL-MIX", lines)], [], "t", fetch_items)
    _ws, rows = sheet_values(data, 0)
    skus = {i.sku for i in catalog.ITEMS}
    total = sum(r[3] for r in rows if r[0] in skus and isinstance(r[3], int))
    assert total == 26


# ── Robustness ───────────────────────────────────────────────────────────────

def test_an_unrecognised_product_still_appears_instead_of_vanishing(stickers):
    """It cannot be weighed, so it cannot be packed — but dropping it silently
    would lose a paid line. It must show up, unlabelled and obvious."""
    orders = [order("NL-ODD", [line("", "Definitely Not A Peptide", "10mg", 2)])]
    data = manifest.build_labeling_manifest(orders, [], "today", fetch_items)
    _ws, rows = sheet_values(data, 0)
    flat = " ".join(str(c) for r in rows for c in r if c)
    assert "NL-ODD" in flat
    assert "Definitely Not A Peptide" in flat
    assert "NO STICKER ON FILE" in flat


def test_an_order_with_no_items_does_not_crash_the_manifest(stickers):
    data = manifest.build_labeling_manifest([order("NL-EMPTY", [])], [], "t", fetch_items)
    _ws, rows = sheet_values(data, 0)
    assert any("NL-EMPTY" in str(c) for r in rows for c in r if c)


def test_the_sheet_is_set_up_to_print(stickers):
    """It gets printed and taped up at a bench. Without fit-to-width the sticker
    column — the one the crew most needs — falls off the right-hand edge."""
    data = manifest.build_labeling_manifest(
        [order("NL-1", [line("RT10", "Retatrutide", "10mg x10", 1)])], [], "t", fetch_items)
    wb = load_workbook(io.BytesIO(data))
    for ws in wb:
        assert ws.page_setup.orientation == "landscape"
        assert ws.sheet_properties.pageSetUpPr.fitToPage is True
        assert ws.print_title_rows.replace("$", "") == "1:3"


def test_customer_address_appears_once_per_order_not_on_every_row(stickers):
    """Jason ships from this sheet, so the address has to be on it — but repeating
    it against every SKU is what made the old manifest unreadable."""
    orders = [order("NL-1", [line("RT10", "Retatrutide", "10mg x10", 1),
                             line("RT100", "Retatrutide", "100mg x10", 1)])]
    data = manifest.build_labeling_manifest(orders, [], "today", fetch_items)
    _ws, rows = sheet_values(data, 0)
    hits = [r for r in rows if r[0] and "12 Main St" in str(r[0])]
    assert len(hits) == 1
