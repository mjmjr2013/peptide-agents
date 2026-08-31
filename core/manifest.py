from __future__ import annotations
"""
The labeling & shipping manifest Jason's crew works from.

WHY THIS EXISTS (2026-08-31). The old warehouse manifest was one flat sheet with
the whole order crushed into a single "Items" cell:

    NL-20260830-1A2B | Jane Doe | 12 Main St… | 3x Retatrutide 10mg x10; 2x Retatrutide 100mg x10

A crew reading that has to parse a semicolon-separated string, work out which
sticker goes on which vial, and keep 10mg and 100mg straight by eye — and
mislabelled strength is the exact failure Northline most wants engineered out.

This builds a workbook shaped the way the crew actually works: ONE ROW PER SKU,
carrying the SKU, the product name exactly as printed on the sticker, the
strength on its own, the number of kits, and a picture of the sticker itself.
Format follows the sheet Daniel produced, which Jason's crew already reads.

TWO TABS, because the two jobs are different and mixing them is what causes
things to be missed:
  1. LABEL & SHIP — orders with no tracking number yet. The crew labels these.
  2. PHOTO BEFORE SHIP — tracking already entered; all that is outstanding is the
     photo of the packed vials, taken right before dispatch (§16/§18a).

PACKAGES ARE BROKEN OUT. A heavy order ships as several parcels under the 2 kg
cap (core/shipping.py), and the crew packs parcels, not orders. Each package gets
its own block with its own contents, so nobody has to divide an order in their
head at the bench.

Nothing here decides anything: it renders what core/catalog and core/shipping
already know. A SKU with no sticker on file is called out IN RED rather than left
blank — a missing label must stop someone, not be quietly skipped.
"""
import io
import re

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

from core import catalog, shipping

# ── Look and feel ────────────────────────────────────────────────────────────
NAVY = "1F2A44"        # Northline label navy
RED = "C0392B"
LIGHT = "EEF1F6"
BAND = "DCE3EF"

_TITLE = Font(bold=True, size=14, color="FFFFFF")
_HEAD = Font(bold=True, size=11, color="FFFFFF")
_ORDER = Font(bold=True, size=12, color="FFFFFF")
_PKG = Font(bold=True, size=11, color=NAVY)
_MISSING = Font(bold=True, size=11, color=RED)
_STRENGTH = Font(bold=True, size=14, color=NAVY)

_THIN = Side(style="thin", color="B7C0D0")
_BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

COLUMNS = [
    ("SKU", 12.5),
    ("Product — as printed on the sticker", 38.0),
    ("Strength", 13.0),
    ("Kits", 8.0),
    ("Sticker Label", 54.0),
]

# Daniel's sheet used 108.75pt rows, but at that height only three rows fit on a
# landscape page and a four-line order spilled onto a page with no header on it.
# 84pt fits five and keeps the printed strength perfectly legible.
LABEL_ROW_HEIGHT = 84.0
LABEL_MAX_PX = (300, 108)   # sticker box inside the cell


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _strength_of(item, spec: str) -> str:
    """The dose on its own, big and bold in its own column.

    The whole point: '10mg' and '100mg' are one character apart in a product
    name and impossible to confuse when they sit alone in a column of their own.
    """
    if item is not None and item.dose is not None:
        dose = int(item.dose) if float(item.dose).is_integer() else item.dose
        return f"{dose}{item.unit}"
    m = re.search(r"\d+\.?\d*\s*(?:mg|ml|iu|mcg|g)", (spec or "").lower())
    return m.group().replace(" ", "") if m else ""


def _address(f: dict) -> str:
    parts = [f.get("address_line1"), f.get("address_line2"),
             " ".join(x for x in [f.get("city"), f.get("state_province"),
                                  f.get("postal_code")] if x),
             f.get("country")]
    return ", ".join(x for x in parts if x)


def _order_items(order: dict, fetch_items) -> list[dict]:
    """Airtable order items → the shape core.shipping speaks."""
    out = []
    for it in fetch_items(order):
        f = it.get("fields", {})
        kits = int(f.get("kits") or 0)
        if kits <= 0:
            continue
        out.append({"sku": f.get("supplier_sku") or "", "product": f.get("product", ""),
                    "spec": f.get("spec", ""), "kits": kits})
    return out


class _SheetWriter:
    def __init__(self, ws, title: str, subtitle: str):
        self.ws = ws
        self.row = 1
        self._images = []          # keep refs alive until save
        ws.freeze_panes = "A4"
        last = get_column_letter(len(COLUMNS))

        # This gets PRINTED and taped up at a bench, so set it up to print. Without
        # this the sticker column falls off the right-hand edge of the page — the
        # one column the crew most needs.
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_title_rows = "1:3"          # headers repeat on every page
        ws.print_options.horizontalCentered = True
        ws.page_margins.left = ws.page_margins.right = 0.3
        ws.page_margins.top = ws.page_margins.bottom = 0.4

        ws.merge_cells(f"A1:{last}1")
        c = ws["A1"]; c.value = title; c.font = _TITLE; c.fill = _fill(NAVY)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 26

        ws.merge_cells(f"A2:{last}2")
        c = ws["A2"]; c.value = subtitle
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 30

        for i, (name, width) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=3, column=i, value=name)
            cell.font = _HEAD; cell.fill = _fill(NAVY)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.row_dimensions[3].height = 30
        self.row = 4

    def _banner(self, text: str, font: Font, color: str, height: int) -> None:
        last = get_column_letter(len(COLUMNS))
        self.ws.merge_cells(f"A{self.row}:{last}{self.row}")
        c = self.ws.cell(row=self.row, column=1, value=text)
        c.font = font; c.fill = _fill(color)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        self.ws.row_dimensions[self.row].height = height
        self.row += 1

    def order_header(self, order_ref: str, name: str, addr: str, phone: str,
                     packages: int) -> None:
        pkg = f"  ·  {packages} package{'s' if packages != 1 else ''}" if packages else ""
        self._banner(f"{order_ref}  —  {name}{pkg}", _ORDER, NAVY, 22)
        self._banner(f"    {addr}" + (f"   ☎ {phone}" if phone else ""),
                     Font(size=10, color=NAVY), LIGHT, 20)

    def package_header(self, order_ref: str, index: int, of: int, kits: int,
                       gross_g: float, uncapped: bool) -> None:
        # The order ref is repeated on EVERY package line on purpose. A big order
        # runs over a page break, and a package header stranded at the top of a
        # fresh page with no order above it is a parcel nobody can identify.
        note = "  (water only — ships whole, no weight cap)" if uncapped else ""
        self._banner(
            f"    {order_ref}  ·  PACKAGE {index} of {of}  —  "
            f"{kits} kit{'s' if kits != 1 else ''}  —  {gross_g / 1000:.2f} kg{note}",
            _PKG, BAND, 20)

    def page_break(self) -> None:
        """Start the next order on a fresh page, so an order's rows stay together."""
        if self.row > 4:
            self.ws.row_breaks.append(Break(id=self.row - 1))

    def item_row(self, sku: str, label_text: str, strength: str, kits: int,
                 image_path) -> None:
        ws, r = self.ws, self.row
        ws.cell(row=r, column=1, value=sku).font = Font(bold=True, size=11)
        ws.cell(row=r, column=2, value=label_text)
        s = ws.cell(row=r, column=3, value=strength); s.font = _STRENGTH
        k = ws.cell(row=r, column=4, value=kits); k.font = Font(bold=True, size=12)
        for col in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=r, column=col)
            cell.border = _BOX
            if col in (3, 4):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center",
                                           wrap_text=True)

        if image_path is not None:
            img = XLImage(str(image_path))
            scale = min(LABEL_MAX_PX[0] / img.width, LABEL_MAX_PX[1] / img.height, 1.0)
            img.width = int(img.width * scale)
            img.height = int(img.height * scale)
            img.anchor = f"{get_column_letter(len(COLUMNS))}{r}"
            ws.add_image(img)
            self._images.append(img)
            ws.row_dimensions[r].height = LABEL_ROW_HEIGHT
        else:
            c = ws.cell(row=r, column=len(COLUMNS),
                        value="⚠ NO STICKER ON FILE — do not label, ask Jordan")
            c.font = _MISSING
            ws.row_dimensions[r].height = 24
        self.row += 1

    def spacer(self) -> None:
        self.ws.row_dimensions[self.row].height = 8
        self.row += 1

    def empty_note(self, text: str) -> None:
        self._banner(text, Font(italic=True, size=11, color=NAVY), LIGHT, 22)


def _write_orders(ws, title: str, subtitle: str, orders: list[dict],
                  fetch_items, empty_note: str) -> None:
    w = _SheetWriter(ws, title, subtitle)
    if not orders:
        w.empty_note(empty_note)
        return

    for order in sorted(orders, key=lambda o: o["fields"].get("order_ref", "")):
        f = order["fields"]
        items = _order_items(order, fetch_items)
        try:
            packages = shipping.split_packages(items)
        except Exception as e:          # never let one odd order kill the manifest
            print(f"[manifest] split failed for {f.get('order_ref')}: {e!r}")
            packages = []

        w.page_break()
        w.order_header(f.get("order_ref", ""), f.get("ship_name", ""),
                       _address(f), f.get("ship_phone", ""), len(packages))

        if not packages:
            # Unresolvable SKUs, or nothing to weigh. Show the raw lines anyway —
            # the crew must see that this order exists, not silently lose it.
            w.package_header(f.get("order_ref", ""), 1, 1,
                             sum(i["kits"] for i in items), 0.0, False)
            for i in items:
                item = catalog.find(i["product"], i["spec"])
                w.item_row(i["sku"] or (item.sku if item else "?"),
                           _label_text(item, i), _strength_of(item, i["spec"]),
                           i["kits"], _label_image(item))
            w.spacer()
            continue

        for n, pkg in enumerate(packages):
            # Break before every package except the first — the first sits under
            # the order header it belongs to, and breaking there would strand
            # that header alone at the foot of the previous page.
            if n:
                w.page_break()
            w.package_header(f.get("order_ref", ""), pkg["index"], pkg["of"],
                             pkg["kits"], pkg["gross_g"], not pkg["capped"])
            for line in pkg["contents"]:
                item = catalog.get(line["sku"])
                w.item_row(line["sku"], _label_text(item, line),
                           _strength_of(item, line.get("spec", "")),
                           line["kits"], _label_image(item))
        w.spacer()


def _label_text(item, line: dict) -> str:
    """What is printed on the sticker, which is what the crew matches against.

    Falls back to the catalog name when no sticker text is on file, so the row is
    still usable — the picture column is where a missing label is called out.
    """
    if item is not None:
        text = catalog.SKU_LABEL_TEXT.get(item.sku)
        if text:
            return text
        return f"{item.product} {_strength_of(item, item.spec)}".strip()
    return f"{line.get('product', '')} {line.get('spec', '')}".strip() or "UNKNOWN PRODUCT"


def _label_image(item):
    if item is None:
        return None
    path = catalog.label_path(item.sku)
    return path if path is not None and path.is_file() else None


def build_labeling_manifest(new_orders: list[dict], photo_orders: list[dict],
                            label: str, fetch_items) -> bytes:
    """The workbook attached to Jason's daily manifest email.

    `fetch_items(order)` returns that order's Airtable Order Item records — passed
    in rather than imported so this module stays testable without Airtable.
    """
    wb = Workbook()

    _write_orders(
        wb.active if wb.active is not None else wb.create_sheet(),
        f"1 · LABEL & SHIP — {label}",
        "New orders. Label every vial, then enter the tracking number on the manifest "
        "page. Check the STRENGTH column against the sticker before you label — that is "
        "the one thing we cannot fix after it ships.",
        new_orders, fetch_items,
        "Nothing new to label today.")
    wb.active.title = "1 · Label & Ship"

    ws2 = wb.create_sheet("2 · Photo Before Ship")
    _write_orders(
        ws2,
        f"2 · PHOTO BEFORE SHIP — {label}",
        "Already labelled and tracked. Photograph the packed vials with the customer's "
        "name and address visible, upload it on the manifest page, and send. Nothing to "
        "re-label here.",
        photo_orders, fetch_items,
        "No orders are waiting on a photo.")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def split_by_stage(orders: list[dict]) -> tuple[list[dict], list[dict]]:
    """Split the outstanding orders into the two tabs.

    Tab 1 is anything without a tracking number — that is the labelling work.
    Tab 2 is tracked already and waiting only on the vial photo. The two are
    disjoint and together they are exactly `get_orders_needing_fulfillment()`.
    """
    new_orders = [o for o in orders if not o["fields"].get("tracking_sent")]
    photo_orders = [o for o in orders
                    if o["fields"].get("tracking_sent")
                    and not o["fields"].get("vial_photo_sent")]
    return new_orders, photo_orders


# ── The view the WEB PAGE renders ────────────────────────────────────────────
# Jordan, 2026-08-31: the manifest page itself should look like this workbook —
# collapsed orders that expand, packages broken out, the sticker beside each row
# — and tracking should be entered there, per package, feeding Airtable directly.
# A spreadsheet mailed alongside is siloed: anything typed into it goes nowhere.
#
# So the page and the workbook are built from the SAME structure below. Neither
# can drift into showing a different picture of an order from the other.

_TRACK_RE = re.compile(r"(\d+)\s*/\s*(\d+)\s+(\S+)")


def parse_tracking(raw: str) -> dict[int, str]:
    """Read the per-package tracking numbers out of the order's tracking field.

    Stored as "1/3 ABC123 | 2/3 DEF456" when an order ships in several packages,
    and as a bare number when it is a single parcel — which is exactly what the
    field held before packages existed, so old orders keep reading correctly.
    """
    raw = (raw or "").strip()
    if not raw:
        return {}
    found = {int(i): num for i, _of, num in _TRACK_RE.findall(raw)}
    return found or {1: raw}


def format_tracking(numbers: dict[int, str], total: int) -> str:
    """The inverse. A single-package order stores the bare number, so nothing
    downstream has to learn a new format for the common case."""
    clean = {int(i): (n or "").strip() for i, n in numbers.items() if (n or "").strip()}
    if not clean:
        return ""
    if total <= 1:
        return clean[min(clean)]
    return " | ".join(f"{i}/{total} {clean[i]}" for i in sorted(clean))


def tracking_numbers(raw: str) -> list[str]:
    """Just the numbers, in package order — what the customer is told."""
    return [n for _i, n in sorted(parse_tracking(raw).items())]


# ── How old is this order? ───────────────────────────────────────────────────
# Jordan, 2026-08-31: the warehouse must work OLDEST FIRST. The worry is concrete
# — a delivery of fresh stock arrives and Jason fulfils whatever is in front of
# him, so an order that has already waited two weeks waits longer still. Sorting
# by age and showing it on every row makes the queue self-policing.
#
# The date is NOT `created_at`: that field exists in the Airtable schema but
# nothing ever writes it, so it is empty on every order. `paid_at` is the right
# business event anyway (the customer's clock starts when their money arrives),
# with Airtable's own createdTime and the date inside the order ref as fallbacks.
_REF_DATE = re.compile(r"[A-Z]+-(\d{4})(\d{2})(\d{2})")


def order_date(order: dict) -> tuple[str, str]:
    """(ISO date string, where it came from). Empty string if genuinely unknown."""
    f = order.get("fields", {})
    paid = (f.get("paid_at") or "").strip()
    if paid:
        return paid[:10], "paid"
    created = (order.get("createdTime") or "").strip()
    if created:
        return created[:10], "created"
    m = _REF_DATE.match((f.get("order_ref") or "").upper())
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}", "ref"
    return "", "unknown"


def days_waiting(iso_date: str, today=None) -> int | None:
    if not iso_date:
        return None
    from datetime import date
    try:
        y, m, d = (int(x) for x in iso_date.split("-")[:3])
    except (ValueError, TypeError):
        return None
    return max(0, ((today or date.today()) - date(y, m, d)).days)


def pretty_date(iso_date: str) -> str:
    """'2026-08-12' -> '12 Aug'. Unambiguous for a crew reading it in China, where
    12/08 and 08/12 mean opposite things."""
    if not iso_date:
        return ""
    from datetime import date
    try:
        y, m, d = (int(x) for x in iso_date.split("-")[:3])
        return date(y, m, d).strftime("%-d %b")
    except (ValueError, TypeError):
        return iso_date


def order_view(order: dict, fetch_items) -> dict:
    """Everything the manifest needs about one order, packages resolved.

    `label_url` points at Flask's static route, so the page shows the same
    sticker image the workbook embeds — one source, two renderings.
    """
    f = order.get("fields", {})
    items = _order_items(order, fetch_items)
    try:
        packages = shipping.split_packages(items)
    except Exception as e:
        print(f"[manifest] split failed for {f.get('order_ref')}: {e!r}")
        packages = []

    def row(sku, product, spec, kits):
        item = catalog.get(sku) if sku else None
        if item is None:
            item = catalog.find(product, spec)
        path = _label_image(item)
        return {
            "sku": (item.sku if item else (sku or "?")),
            "label_text": _label_text(item, {"product": product, "spec": spec}),
            "strength": _strength_of(item, spec or (item.spec if item else "")),
            "kits": kits,
            "label_url": (f"/static/labels/{path.name}" if path else None),
        }

    views = []
    if packages:
        for pkg in packages:
            views.append({
                "index": pkg["index"], "of": pkg["of"], "kits": pkg["kits"],
                "gross_g": pkg["gross_g"], "capped": pkg["capped"],
                "rows": [row(c["sku"], c["product"], c.get("spec", ""), c["kits"])
                         for c in pkg["contents"]],
            })
    elif items:
        # Unresolvable products cannot be weighed, so they cannot be packed — but
        # they must still be visible, or a paid line silently vanishes.
        views.append({
            "index": 1, "of": 1, "kits": sum(i["kits"] for i in items),
            "gross_g": 0.0, "capped": True,
            "rows": [row(i["sku"], i["product"], i["spec"], i["kits"]) for i in items],
        })

    ordered_at, date_source = order_date(order)
    tracking = parse_tracking(f.get("tracking_number", ""))
    total = len(views) or 1
    return {
        "id": order.get("id", ""),
        "ref": f.get("order_ref", "") or order.get("id", ""),
        "name": f.get("ship_name", ""),
        "address": _address(f),
        "phone": f.get("ship_phone", ""),
        "kits": sum(p["kits"] for p in views),
        "ordered_at": ordered_at,
        "ordered_label": pretty_date(ordered_at),
        "date_source": date_source,
        "age_days": days_waiting(ordered_at),
        "packages": views,
        "package_count": total,
        "tracking": tracking,
        # Complete only when EVERY package has a number. Until then the order
        # stays on the list and the customer is not messaged — a buyer told
        # "shipped" while two of three parcels have no label is worse than
        # waiting (HANDOFF §18a: tracking means "booked", not "gone").
        "tracking_complete": all(tracking.get(i) for i in range(1, total + 1)),
        "vial_photo_sent": bool(f.get("vial_photo_sent")),
        "tracking_sent": bool(f.get("tracking_sent")),
    }


def build_view(orders: list[dict], fetch_items) -> list[dict]:
    """OLDEST FIRST — the order the warehouse should work in.

    An order with no resolvable date sorts last rather than first: every order on
    this page is paid, so a date should always exist, and one that is missing is
    an oddity that should not be able to jump the queue. The page badges it.
    """
    views = [order_view(o, fetch_items) for o in orders]
    return sorted(views, key=lambda v: (v["ordered_at"] == "", v["ordered_at"],
                                        v["ref"]))
