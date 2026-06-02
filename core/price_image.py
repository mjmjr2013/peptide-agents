"""
Generates a price list image (PNG) for WhatsApp delivery.
2-column layout, light theme, maximum readability.
"""
from __future__ import annotations
from pathlib import Path

import os as _os
_STATIC = Path(__file__).parent.parent / "static"
_ICLOUD = Path.home() / "Library" / "Mobile Documents" / "com~apple~CloudDocs" / "Documents" / "Northline Group"

# On Railway (or any non-Mac env) save to static/; locally save to iCloud folder
if _os.environ.get("RAILWAY_ENVIRONMENT") or not (_ICLOUD.parent.exists()):
    OUTPUT_PATH    = _STATIC / "price_list.png"
    CN_OUTPUT_PATH = _STATIC / "price_list_cn.png"
    XLSX_PATH      = _STATIC / "price_list.xlsx"
    PDF_PATH       = _STATIC / "price_list.pdf"
else:
    OUTPUT_PATH    = _ICLOUD / "price_list.png"
    CN_OUTPUT_PATH = _ICLOUD / "price_list_cn.png"
    XLSX_PATH      = _ICLOUD / "price_list.xlsx"
    PDF_PATH       = _ICLOUD / "price_list.pdf"

CATEGORIES = [
    ("GLP-1 Peptides", [
        ("SM5",    "Semaglutide",           "5mg",      "$58"),
        ("SM10",   "Semaglutide",           "10mg",     "$92"),
        ("SM15",   "Semaglutide",           "15mg",     "$133"),
        ("SM20",   "Semaglutide",           "20mg",     "$166"),
        ("SM30",   "Semaglutide",           "30mg",     "$216"),
        ("SM40",   "Semaglutide",           "40mg",     "$288"),
        ("SM50",   "Semaglutide",           "50mg",     "$360"),
        ("TR5",    "Tirzepatide",           "5mg",      "$71"),
        ("TR10",   "Tirzepatide",           "10mg",     "$108"),
        ("TR15",   "Tirzepatide",           "15mg",     "$162"),
        ("TR20",   "Tirzepatide",           "20mg",     "$216"),
        ("TR30",   "Tirzepatide",           "30mg",     "$274"),
        ("TR40",   "Tirzepatide",           "40mg",     "$365"),
        ("TR50",   "Tirzepatide",           "50mg",     "$456"),
        ("TR60",   "Tirzepatide",           "60mg",     "$547"),
        ("TR80",   "Tirzepatide",           "80mg",     "$729"),
        ("TR100",  "Tirzepatide",           "100mg",    "$795"),
        ("RT5",    "Retatrutide",           "5mg",      "$67"),
        ("RT10",   "Retatrutide",           "10mg",     "$95"),
        ("RT15",   "Retatrutide",           "15mg",     "$142"),
        ("RT20",   "Retatrutide",           "20mg",     "$189"),
        ("RT30",   "Retatrutide",           "30mg",     "$274"),
        ("RT40",   "Retatrutide",           "40mg",     "$365"),
        ("RT50",   "Retatrutide",           "50mg",     "$456"),
        ("RT60",   "Retatrutide",           "60mg",     "$547"),
        ("RT80",   "Retatrutide",           "80mg",     "$729"),
        ("RT100",  "Retatrutide",           "100mg",    "$894"),
        ("CGL5",   "Cagrilintide",          "5mg",      "$115"),
        ("CGL10",  "Cagrilintide",          "10mg",     "$181"),
        ("MDT5",   "Mazdutide",             "5mg",      "$192"),
        ("MDT10",  "Mazdutide",             "10mg",     "$203"),
        ("DUL5",   "Dulaglutide",           "5mg",      "$315"),
        ("DUL10",  "Dulaglutide",           "10mg",     "$514"),
        ("SUR2",   "Survodutide",           "2mg",      "$265"),
        ("SUR5",   "Survodutide",           "5mg",      "$480"),
        ("SUR10",  "Survodutide",           "10mg",     "$820"),
        ("LGT5",   "Liraglutide",           "5mg",      "$224"),
        ("LGT10",  "Liraglutide",           "10mg",     "$398"),
        ("LGT20",  "Liraglutide",           "20mg",     "$737"),
    ]),
    ("Healing & Recovery", [
        ("BC5",    "BPC-157",               "5mg",      "$58"),
        ("BC10",   "BPC-157",               "10mg",     "$72"),
        ("BT5",    "TB-500",                "5mg",      "$92"),
        ("BT10",   "TB-500",                "10mg",     "$140"),
        ("BB10",   "BPC+TB Blend",          "10mg",     "$108"),
        ("BB20",   "BPC+TB Blend",          "20mg",     "$166"),
        ("GLOW70", "BPC+TB+GHK Blend",      "70mg",     "$154"),
        ("KLOW",   "BPC+TB+GHK+KPV",        "80mg",     "$220"),
        ("CU50",   "GHK-Cu",                "50mg",     "$71"),
        ("CU100",  "GHK-Cu",                "100mg",    "$116"),
        ("KPV5",   "KPV",                   "5mg",      "$63"),
        ("KPV10",  "KPV",                   "10mg",     "$100"),
        ("P41",    "PT-141",                "10mg",     "$72"),
        ("ML10",   "Melanotan II",          "10mg",     "$149"),
        ("2AD",    "AOD-9604",              "2mg",      "$100"),
        ("5AD",    "AOD-9604",              "5mg",      "$191"),
        ("10AD",   "AOD-9604",              "10mg",     "$332"),
        ("GND2",   "Gonadorelin",           "2mg",      "$56"),
        ("AP2",    "Adipotide",             "2mg",      "$86"),
        ("AP5",    "Adipotide",             "5mg",      "$166"),
        ("RA10",   "Ara-290",               "10mg",     "$149"),
        ("RA16",   "Ara-290",               "16mg",     "$238"),
        ("DR2",    "Dermorphin",            "2mg",      "$72"),
        ("DR5",    "Dermorphin",            "5mg",      "$125"),
        ("DR10",   "Dermorphin",            "10mg",     "$199"),
        ("DR20",   "Dermorphin",            "20mg",     "$332"),
        ("NP810",  "Snap-8",                "10mg",     "$133"),
        ("NP8100", "Snap-8",                "100mg",    "$663"),
        ("LC216",  "Lipo-C",                "10ml",     "$92"),
        ("MIC10",  "MIC (Lipo+B12)",        "10ml",     "$298"),
    ]),
    ("GH / Growth", [
        ("IP2",    "Ipamorelin",            "2mg",      "$47"),
        ("IP5",    "Ipamorelin",            "5mg",      "$58"),
        ("IP10",   "Ipamorelin",            "10mg",     "$100"),
        ("CND2",   "CJC-1295 (no DAC)",     "2mg",      "$42"),
        ("CND5",   "CJC-1295 (no DAC)",     "5mg",      "$98"),
        ("CND10",  "CJC-1295 (no DAC)",     "10mg",     "$158"),
        ("CD5",    "CJC-1295 (w/ DAC)",     "5mg",      "$166"),
        ("CP10",   "CJC+Ipa Blend",         "10mg",     "$114"),
        ("SMO5",   "Sermorelin",            "5mg",      "$90"),
        ("SMO10",  "Sermorelin",            "10mg",     "$119"),
        ("G25",    "GHRP-2",               "5mg",      "$34"),
        ("G210",   "GHRP-2",               "10mg",     "$58"),
        ("G65",    "GHRP-6",               "5mg",      "$38"),
        ("G610",   "GHRP-6",               "10mg",     "$42"),
        ("HX2",    "Hexarelin",             "2mg",      "$56"),
        ("HX5",    "Hexarelin",             "5mg",      "$104"),
        ("KS5",    "KissPeptin-10",         "5mg",      "$72"),
        ("KS10",   "KissPeptin-10",         "10mg",     "$116"),
        ("FM2",    "MGF",                   "2mg",      "$58"),
        ("FMP2",   "PEG MGF",              "2mg",      "$101"),
        ("H8",     "HGH 191AA",             "8iu",      "$65"),
        ("H10",    "HGH 191AA",             "10iu",     "$80"),
        ("H15",    "HGH 191AA",             "15iu",     "$106"),
        ("G5K",    "HCG",                   "5000IU",   "$104"),
        ("G10K",   "HCG",                   "10000IU",  "$164"),
        ("IG1",    "IGF-1 LR3",             "1mg",      "$204"),
        ("IGD",    "IGF-DES",               "2mg",      "$77"),
        ("FN1",    "Follistatin",           "1mg",      "$290"),
        ("AE1",    "ACE-031",               "1mg",      "$243"),
        ("TY10",   "Thymalin",              "10mg",     "$77"),
        ("TSM2",   "Tesamorelin",           "2mg",      "$72"),
        ("TSM5",   "Tesamorelin",           "5mg",      "$115"),
        ("TSM10",  "Tesamorelin",           "10mg",     "$195"),
        ("TSM20",  "Tesamorelin",           "20mg",     "$290"),
        ("ACTH5",  "ACTH",                  "5mg",      "$183"),
        ("EP0",    "EPO",                   "3000IU",   "$149"),
    ]),
    ("Cognitive & Wellness", [
        ("NJ100",  "NAD",                   "100mg",    "$100"),
        ("NJ500",  "NAD",                   "500mg",    "$240"),
        ("NJ1000", "NAD",                   "1000mg",   "$265"),
        ("GTT4",   "Glutathione",           "400mg",    "$67"),
        ("GTT",    "Glutathione",           "600mg",    "$87"),
        ("GTT15",  "Glutathione",           "1500mg",   "$166"),
        ("ET10",   "Epithalon",             "10mg",     "$64"),
        ("ET50",   "Epithalon",             "50mg",     "$240"),
        ("MS10",   "MOTS-c",                "10mg",     "$82"),
        ("MS20",   "MOTS-c",                "20mg",     "$112"),
        ("MS40",   "MOTS-c",                "40mg",     "$197"),
        ("TA2",    "Thymosin Alpha-1",      "2mg",      "$73"),
        ("TA5",    "Thymosin Alpha-1",      "5mg",      "$105"),
        ("TA10",   "Thymosin Alpha-1",      "10mg",     "$176"),
        ("5AM",    "5-Amino/MQ",            "5mg",      "$183"),
        ("5AM10",  "5-Amino/MQ",            "10mg",     "$261"),
        ("50AM",   "5-Amino/MQ",            "50mg",     "$812"),
        ("SK5",    "Selank",                "5mg",      "$55"),
        ("SK10",   "Selank",                "10mg",     "$92"),
        ("XA5",    "Semax",                 "5mg",      "$53"),
        ("XA10",   "Semax",                 "10mg",     "$92"),
        ("DS2",    "DSIP",                  "2mg",      "$38"),
        ("DS5",    "DSIP",                  "5mg",      "$58"),
        ("DS10",   "DSIP",                  "10mg",     "$104"),
        ("PI5",    "Pinealon",              "5mg",      "$75"),
        ("PI10",   "Pinealon",              "10mg",     "$125"),
        ("MEL10",  "Melatonin",             "10mg",     "$133"),
        ("OT2",    "Oxytocin",              "2mg",      "$72"),
        ("OT5",    "Oxytocin",              "5mg",      "$125"),
        ("OT10",   "Oxytocin",              "10mg",     "$232"),
        ("AR50",   "AICAR",                 "50mg",     "$80"),
        ("2S10",   "SS-31",                 "10mg",     "$100"),
        ("2S50",   "SS-31",                 "50mg",     "$414"),
        ("ADA5",   "Admax",                 "5mg",      "$158"),
        ("ADA10",  "Admax",                 "10mg",     "$265"),
        ("F42",    "FOXO4-DRI",             "2mg",      "$232"),
        ("F45",    "FOXO4-DRI",             "5mg",      "$373"),
        ("F410",   "FOXO4-DRI",             "10mg",     "$629"),
        ("CAR10",  "Cardiogen",             "10mg",     "$174"),
        ("CAR20",  "Cardiogen",             "20mg",     "$298"),
        ("CART10", "Cartalax",              "10mg",     "$191"),
        ("CART20", "Cartalax",              "20mg",     "$323"),
        ("CRY10",  "Crystagen",             "10mg",     "$158"),
        ("CRY20",  "Crystagen",             "20mg",     "$290"),
        ("HUM10",  "Humanin",               "10mg",     "$737"),
        ("MAT10",  "Matrixyl",              "10mg",     "$82"),
        ("PN5",    "PNC-27",                "5mg",      "$290"),
        ("SLU5",   "SLU-PP-322",            "5mg",      "$216"),
    ]),
]


def generate_price_list_image(lang: str = "en") -> Path:
    import matplotlib
    matplotlib.use("Agg")

    # ── Language / text setup ─────────────────────────────────────────────────
    if lang == "cn":
        import matplotlib as mpl
        mpl.rcParams["font.family"] = ["Songti SC", "PingFang HK",
                                       "Hiragino Sans GB", "Arial Unicode MS",
                                       "DejaVu Sans"]
        out_path   = CN_OUTPUT_PATH
        title_text = "北线集团 — 研究肽价格表  (NORTHLINE GROUP — RESEARCH PEPTIDE PRICE LIST)"
        sub_text   = "每套 (10瓶)  •  所有价格USD  •  批量优惠可谈  (Per Kit • All Prices USD • Volume Pricing Available)"
        footer_text = "所有产品仅供研究使用  •  最低订购: 1套  •  价格可能变动  (Research use only • Min order: 1 kit • Prices subject to change)"
        hdr_sku    = "SKU"
        hdr_prod   = "产品 (PRODUCT)"
        hdr_spec   = "规格 (SPEC)"
        hdr_price  = "价格/套 (PRICE/KIT)"
        cat_map    = {
            "GLP-1 Peptides":      "GLP-1 肽类  (GLP-1 Peptides)",
            "Healing & Recovery":  "愈合恢复  (Healing & Recovery)",
            "GH / Growth":         "生长激素  (GH / Growth)",
            "Cognitive & Wellness":"认知健康  (Cognitive & Wellness)",
        }
    else:
        out_path    = OUTPUT_PATH
        title_text  = "NORTHLINE GROUP  —  RESEARCH PEPTIDE PRICE LIST"
        sub_text    = "Per Kit (10 Vials)  •  All Prices USD  •  Volume Pricing Available"
        footer_text = "All products for research use only  •  Minimum order: 1 kit  •  Prices subject to change"
        hdr_sku   = "SKU"
        hdr_prod  = "PRODUCT"
        hdr_spec  = "SPEC"
        hdr_price = "PRICE / KIT"
        cat_map   = {}
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches

    # ── Palette ──────────────────────────────────────────────────────────────
    BG          = "#FFFFFF"
    HEADER_BG   = "#2C3E50"
    CAT_BG      = "#BDC3C7"
    CAT_TEXT    = "#1C1C1E"
    BORDER      = "#D0D0D0"
    TEXT_MAIN   = "#1C1C1E"
    TEXT_DIM    = "#555555"
    PRICE_COLOR = "#1C1C1E"
    COL_HEAD    = "#FFFFFF"

    # Per-product color bands (light pastels, cycling through products)
    PRODUCT_COLORS = [
        "#FFF9C4",  # soft yellow
        "#C8E6C9",  # soft green
        "#BBDEFB",  # soft blue
        "#F8BBD0",  # soft pink
        "#FFE0B2",  # soft orange
        "#E1BEE7",  # soft purple
        "#B2EBF2",  # soft cyan
        "#DCEDC8",  # soft lime
        "#FFCCBC",  # soft deep orange
        "#CFD8DC",  # soft blue-grey
        "#F0F4C3",  # soft yellow-green
        "#D7CCC8",  # soft brown
        "#B3E5FC",  # light sky blue
        "#F9FBE7",  # cream
        "#FCE4EC",  # light rose
        "#E8F5E9",  # mint
    ]

    # ── Layout ───────────────────────────────────────────────────────────────
    col_gap = 0.04   # gap between 2 columns
    col_w   = (1.0 - col_gap) / 2
    pad     = 0.012

    # Split categories into 2 columns: left = GLP-1 + Healing, right = GH + Wellness
    left_cats  = CATEGORIES[:2]
    right_cats = CATEGORIES[2:]

    def col_rows(cats):
        return sum(len(r) for _, r in cats) + len(cats)

    n_left  = col_rows(left_cats)
    n_right = col_rows(right_cats)
    n_rows  = max(n_left, n_right)

    # Dynamic row/cat sizing so all content always fits in body_h
    # y_frac budget ≤ 1.0:  col_hdr + n_cats*(cat_h + gap) + n_data*row_h
    # With cat_h = 1.5*row_h and gap = 0.4*row_h:
    n_cats_max    = max(len(left_cats), len(right_cats))
    n_product_max = n_rows - n_cats_max
    coeff = 1.5 + n_cats_max * (1.5 + 0.4) + n_product_max
    row_h = min(0.022, 0.93 / coeff)
    cat_h = 1.5 * row_h
    gap   = 0.4 * row_h

    FW = 22
    FH = 2.8 + n_rows * 0.32

    fig = plt.figure(figsize=(FW, FH), facecolor=BG)
    ax  = fig.add_axes([0, 0, 1, 1])
    ax.set_facecolor(BG)
    ax.axis("off")
    ax.set_xlim(0, 1)

    # Reserve space: title at top, footer at bottom
    title_h  = 0.07
    footer_h = 0.035
    body_h   = 1.0 - title_h - footer_h

    def fy(raw_y):
        """Map raw_y (0=top of body … 1=bottom of body) to axes fraction."""
        return (1.0 - title_h) - raw_y * body_h

    # ── Title bar ────────────────────────────────────────────────────────────
    ax.add_patch(mpatches.FancyBboxPatch(
        (0, 1 - title_h), 1, title_h,
        boxstyle="square,pad=0", facecolor=HEADER_BG, edgecolor="none",
        transform=ax.transAxes, clip_on=False))
    ax.text(0.5, 1 - title_h / 2,
            title_text,
            fontsize=24, fontweight="bold", color="white",
            ha="center", va="center", transform=ax.transAxes)
    ax.text(0.5, 1 - title_h + 0.008,
            sub_text,
            fontsize=13, color="#AAAAAA",
            ha="center", va="bottom", transform=ax.transAxes)

    # ── Column header row ────────────────────────────────────────────────────
    def draw_col_headers(x0):
        # Column proportions: SKU(13%) | PRODUCT(39%) | SPEC(22%) | PRICE(26%)
        y0 = fy(0)
        ax.add_patch(plt.Rectangle(
            (x0, y0 - cat_h), col_w, cat_h,
            facecolor=HEADER_BG, edgecolor="none",
            transform=ax.transAxes, clip_on=False))
        ax.text(x0 + 0.006, y0 - cat_h / 2, hdr_sku,
                fontsize=11, fontweight="bold", color=COL_HEAD,
                va="center", transform=ax.transAxes)
        ax.text(x0 + col_w * 0.15, y0 - cat_h / 2, hdr_prod,
                fontsize=11, fontweight="bold", color=COL_HEAD,
                va="center", transform=ax.transAxes)
        ax.text(x0 + col_w * 0.54, y0 - cat_h / 2, hdr_spec,
                fontsize=11, fontweight="bold", color=COL_HEAD,
                va="center", transform=ax.transAxes)
        ax.text(x0 + col_w * 0.87, y0 - cat_h / 2, hdr_price,
                fontsize=11, fontweight="bold", color=COL_HEAD,
                ha="center", va="center", transform=ax.transAxes)

    draw_col_headers(0)
    draw_col_headers(col_w + col_gap)

    # ── Draw one column of categories ────────────────────────────────────────
    def draw_column(cats, x0):
        y_frac = cat_h  # start just below col headers
        color_idx = 0   # global product color counter

        for cat_name, rows in cats:
            y = fy(y_frac)

            # Category header
            ax.add_patch(plt.Rectangle(
                (x0, y - cat_h), col_w, cat_h,
                facecolor=CAT_BG, edgecolor=BORDER, linewidth=0.5,
                transform=ax.transAxes, clip_on=False))
            display_name = cat_map.get(cat_name, cat_name).upper()
            ax.text(x0 + 0.01, y - cat_h / 2, display_name,
                    fontsize=13, fontweight="bold", color=CAT_TEXT,
                    va="center", transform=ax.transAxes)
            y_frac += cat_h

            current_product = None
            row_color = PRODUCT_COLORS[color_idx % len(PRODUCT_COLORS)]

            for i, (sku, product, spec, price) in enumerate(rows):
                y = fy(y_frac)

                # New product = new color band
                if product != current_product:
                    current_product = product
                    color_idx += 1
                    row_color = PRODUCT_COLORS[color_idx % len(PRODUCT_COLORS)]

                # Row background
                ax.add_patch(plt.Rectangle(
                    (x0, y - row_h), col_w, row_h,
                    facecolor=row_color, edgecolor=BORDER, linewidth=0.3,
                    transform=ax.transAxes, clip_on=False))

                # SKU code
                ax.text(x0 + 0.006, y - row_h / 2, sku,
                        fontsize=11, fontweight="bold", color=TEXT_DIM,
                        va="center", transform=ax.transAxes)

                # Product name
                ax.text(x0 + col_w * 0.15, y - row_h / 2, product,
                        fontsize=12, color=TEXT_MAIN, fontweight="bold",
                        va="center", transform=ax.transAxes)

                # Spec
                ax.text(x0 + col_w * 0.54, y - row_h / 2, spec,
                        fontsize=12, color=TEXT_DIM,
                        va="center", transform=ax.transAxes)

                # Price
                ax.text(x0 + col_w * 0.87, y - row_h / 2, price,
                        fontsize=12, fontweight="bold", color=PRICE_COLOR,
                        ha="center", va="center", transform=ax.transAxes)

                y_frac += row_h

            y_frac += gap  # gap after each category

    draw_column(left_cats,  0)
    draw_column(right_cats, col_w + col_gap)

    # ── Footer ───────────────────────────────────────────────────────────────
    ax.text(0.5, footer_h / 2,
            footer_text,
            fontsize=11, color=TEXT_DIM, ha="center", va="center",
            transform=ax.transAxes)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, dpi=250, bbox_inches="tight",
                facecolor=BG, edgecolor="none")
    plt.close(fig)
    print(f"[PriceImage] Saved to {out_path}")
    return out_path


def generate_price_list_image_cn() -> Path:
    """Generate the Chinese/English bilingual price list image."""
    return generate_price_list_image(lang="cn")


def generate_price_list_pdf() -> Path:
    """Generate the bilingual price list as a PDF for WhatsApp delivery."""
    PDF_PATH.parent.mkdir(parents=True, exist_ok=True)
    generate_price_list_image(lang="cn")  # reuse CN render logic to a temp PNG, then save PDF
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches

    BG = "#FFFFFF"; HEADER_BG = "#2C3E50"; CAT_BG = "#BDC3C7"
    CAT_TEXT = "#1C1C1E"; BORDER = "#D0D0D0"; TEXT_MAIN = "#1C1C1E"
    TEXT_DIM = "#555555"; PRICE_COLOR = "#1C1C1E"; COL_HEAD = "#FFFFFF"
    PRODUCT_COLORS = ["#FFF9C4","#C8E6C9","#BBDEFB","#F8BBD0","#FFE0B2",
                      "#E1BEE7","#B2EBF2","#DCEDC8","#FFCCBC","#CFD8DC",
                      "#F0F4C3","#D7CCC8","#B3E5FC","#F9FBE7","#FCE4EC","#E8F5E9"]
    CAT_LABELS = {"GLP-1 Peptides":"GLP-1 肽类  (GLP-1 Peptides)",
                  "Healing & Recovery":"愈合恢复  (Healing & Recovery)",
                  "GH / Growth":"生长激素  (GH / Growth)",
                  "Cognitive & Wellness":"认知健康  (Cognitive & Wellness)"}
    title_text = "NORTHLINE GROUP  —  RESEARCH PEPTIDE PRICE LIST"
    sub_text   = "北线集团 — 每套 (10瓶) · 所有价格USD · 批量优惠可谈  |  Per Kit (10 Vials) · All Prices USD"

    col_gap = 0.04; col_w = (1.0 - col_gap) / 2
    left_cats = CATEGORIES[:2]; right_cats = CATEGORIES[2:]
    def col_rows(cats): return sum(len(r) for _, r in cats) + len(cats)
    n_left = col_rows(left_cats); n_right = col_rows(right_cats)
    n_rows = max(n_left, n_right)
    n_cats_max = max(len(left_cats), len(right_cats))
    n_product_max = n_rows - n_cats_max
    coeff = 1.5 + n_cats_max * (1.5 + 0.4) + n_product_max
    row_h = min(0.022, 0.93 / coeff); cat_h = 1.5 * row_h; gap = 0.4 * row_h
    FW = 22; FH = 2.8 + n_rows * 0.32
    fig = plt.figure(figsize=(FW, FH), facecolor=BG)
    ax = fig.add_axes([0, 0, 1, 1]); ax.set_facecolor(BG); ax.axis("off"); ax.set_xlim(0, 1)
    title_h = 0.07; footer_h = 0.035; body_h = 1.0 - title_h - footer_h
    def fy(raw_y): return (1.0 - title_h) - raw_y * body_h
    ax.add_patch(mpatches.FancyBboxPatch((0, 1-title_h), 1, title_h, boxstyle="square,pad=0",
        facecolor=HEADER_BG, edgecolor="none", transform=ax.transAxes, clip_on=False))
    ax.text(0.5, 1-title_h/2, title_text, fontsize=24, fontweight="bold", color="white",
            ha="center", va="center", transform=ax.transAxes)
    ax.text(0.5, 1-title_h+0.008, sub_text, fontsize=13, color="#AAAAAA",
            ha="center", va="bottom", transform=ax.transAxes)
    def draw_col_headers(x0):
        y0 = fy(0)
        ax.add_patch(plt.Rectangle((x0, y0-cat_h), col_w, cat_h, facecolor=HEADER_BG,
            edgecolor="none", transform=ax.transAxes, clip_on=False))
        for txt, xf in [("SKU / 产品代码", 0.006), ("PRODUCT / 产品", col_w*0.15),
                        ("SPEC / 规格", col_w*0.54), ("PRICE/KIT / 价格", col_w*0.87)]:
            ax.text(x0+xf, y0-cat_h/2, txt, fontsize=11, fontweight="bold", color=COL_HEAD,
                    va="center", ha="center" if "PRICE" in txt else "left", transform=ax.transAxes)
    draw_col_headers(0); draw_col_headers(col_w + col_gap)
    def draw_column(cats, x0):
        y_frac = cat_h; color_idx = 0
        for cat_name, rows in cats:
            y = fy(y_frac)
            ax.add_patch(plt.Rectangle((x0, y-cat_h), col_w, cat_h, facecolor=CAT_BG,
                edgecolor=BORDER, linewidth=0.5, transform=ax.transAxes, clip_on=False))
            ax.text(x0+0.01, y-cat_h/2, CAT_LABELS.get(cat_name, cat_name).upper(),
                    fontsize=13, fontweight="bold", color=CAT_TEXT, va="center", transform=ax.transAxes)
            y_frac += cat_h; current_product = None
            row_color = PRODUCT_COLORS[color_idx % len(PRODUCT_COLORS)]
            for sku, product, spec, price in rows:
                y = fy(y_frac)
                if product != current_product:
                    current_product = product; color_idx += 1
                    row_color = PRODUCT_COLORS[color_idx % len(PRODUCT_COLORS)]
                ax.add_patch(plt.Rectangle((x0, y-row_h), col_w, row_h, facecolor=row_color,
                    edgecolor=BORDER, linewidth=0.3, transform=ax.transAxes, clip_on=False))
                ax.text(x0+0.006, y-row_h/2, sku, fontsize=11, fontweight="bold",
                        color=TEXT_DIM, va="center", transform=ax.transAxes)
                ax.text(x0+col_w*0.15, y-row_h/2, product, fontsize=12, color=TEXT_MAIN,
                        fontweight="bold", va="center", transform=ax.transAxes)
                ax.text(x0+col_w*0.54, y-row_h/2, spec, fontsize=12, color=TEXT_DIM,
                        va="center", transform=ax.transAxes)
                ax.text(x0+col_w*0.87, y-row_h/2, price, fontsize=12, fontweight="bold",
                        color=PRICE_COLOR, ha="center", va="center", transform=ax.transAxes)
                y_frac += row_h
            y_frac += gap
    draw_column(left_cats, 0); draw_column(right_cats, col_w + col_gap)
    ax.text(0.5, footer_h/2,
            "所有产品仅供研究使用 · 最低订购: 1套  |  Research use only · Min order: 1 kit",
            fontsize=11, color=TEXT_DIM, ha="center", va="center", transform=ax.transAxes)
    fig.savefig(PDF_PATH, dpi=150, bbox_inches="tight", facecolor=BG, edgecolor="none")
    plt.close(fig)
    print(f"[PriceImage] Saved PDF to {PDF_PATH}")
    return PDF_PATH


def generate_price_list_xlsx() -> Path:
    """Generate a bilingual xlsx price list for WhatsApp delivery."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    CAT_LABELS = {
        "GLP-1 Peptides":      "GLP-1 肽类  (GLP-1 Peptides)",
        "Healing & Recovery":  "愈合恢复  (Healing & Recovery)",
        "GH / Growth":         "生长激素  (GH / Growth)",
        "Cognitive & Wellness":"认知健康  (Cognitive & Wellness)",
    }
    PRODUCT_BANDS = [
        "FFF9C4", "C8E6C9", "BBDEFB", "F8BBD0", "FFE0B2",
        "E1BEE7", "B2EBF2", "DCEDC8", "FFCCBC", "CFD8DC",
        "F0F4C3", "D7CCC8", "B3E5FC", "F9FBE7", "FCE4EC", "E8F5E9",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Price List"

    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title rows
    ws.merge_cells("A1:D1")
    t1 = ws["A1"]
    t1.value = "NORTHLINE GROUP  —  RESEARCH PEPTIDE PRICE LIST"
    t1.font = Font(bold=True, size=14, color="FFFFFF")
    t1.fill = PatternFill("solid", fgColor="2C3E50")
    t1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    t2 = ws["A2"]
    t2.value = "北线集团 — 每套 (10瓶) · 所有价格USD · 批量优惠可谈  |  Per Kit (10 Vials) · All Prices USD · Volume Pricing Available"
    t2.font = Font(italic=True, size=10, color="555555")
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Column headers
    headers = ["SKU  /  产品代码", "PRODUCT  /  产品", "SPEC  /  规格", "PRICE / KIT  /  价格/套"]
    row = 3
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2C3E50")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[row].height = 20
    row += 1

    color_idx = 0
    for cat_name, items in CATEGORIES:
        # Category header
        ws.merge_cells(f"A{row}:D{row}")
        cat_cell = ws[f"A{row}"]
        cat_cell.value = CAT_LABELS.get(cat_name, cat_name).upper()
        cat_cell.font = Font(bold=True, size=11, color="1C1C1E")
        cat_cell.fill = PatternFill("solid", fgColor="BDC3C7")
        cat_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cat_cell.border = border
        ws.row_dimensions[row].height = 18
        row += 1

        current_product = None
        band = PRODUCT_BANDS[color_idx % len(PRODUCT_BANDS)]

        for sku, product, spec, price in items:
            if product != current_product:
                current_product = product
                color_idx += 1
                band = PRODUCT_BANDS[color_idx % len(PRODUCT_BANDS)]

            fill = PatternFill("solid", fgColor=band.lstrip("#"))
            for col, val in enumerate([sku, product, spec, price], 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(vertical="center", indent=1 if col < 4 else 0,
                                           horizontal="center" if col == 4 else "left")
                if col == 1:
                    cell.font = Font(bold=True, size=9, color="555555")
                elif col == 2:
                    cell.font = Font(bold=True, size=10)
                elif col == 4:
                    cell.font = Font(bold=True, size=10)
                else:
                    cell.font = Font(size=10)
            ws.row_dimensions[row].height = 16
            row += 1

    # Footer
    ws.merge_cells(f"A{row}:D{row}")
    f = ws[f"A{row}"]
    f.value = "所有产品仅供研究使用 · 最低订购: 1套 · 价格可能变动  |  Research use only · Min order: 1 kit · Prices subject to change"
    f.font = Font(italic=True, size=9, color="888888")
    f.alignment = Alignment(horizontal="center")
    ws.row_dimensions[row].height = 16

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16

    ws.freeze_panes = "A4"

    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_PATH)
    print(f"[PriceImage] Saved xlsx to {XLSX_PATH}")
    return XLSX_PATH


if __name__ == "__main__":
    generate_price_list_image()
    generate_price_list_image_cn()
    generate_price_list_xlsx()
    generate_price_list_pdf()
    print("Done.")
